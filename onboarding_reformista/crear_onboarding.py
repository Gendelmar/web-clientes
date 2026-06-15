#!/usr/bin/env python3
"""
Genera el Excel de onboarding para reformistas.
Produce: onboarding_reformista.xlsm (con macro VBA embebida)
Uso: python3 crear_onboarding.py
"""

import os
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paleta de colores ─────────────────────────────────────────
C_BRAND      = "1A1A2E"   # Azul muy oscuro (header principal)
C_ACCENT     = "E94560"   # Rojo/coral (acentos)
C_LIGHT      = "F5F5F5"   # Gris claro (filas pares)
C_WHITE      = "FFFFFF"
C_HEADER_SEC = "16213E"   # Azul oscuro (cabeceras de sección)
C_HEADER_TAB = "0F3460"   # Azul medio (cabeceras de tabla)
C_LABEL      = "2D2D2D"   # Texto oscuro
C_MUTED      = "888888"   # Texto secundario
C_GREEN      = "27AE60"   # Verde OK
C_YELLOW     = "F39C12"   # Amarillo aviso
C_NAV_BG     = "0D0D1A"   # Barra de navegación

# ── Estilos de fuente/relleno reutilizables ───────────────────
def ft(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color=C_ACCENT)
    return Border(bottom=s)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Helpers para escribir celdas ──────────────────────────────
def write(ws, row, col, value, bold=False, size=11, color="000000",
          bg=None, align_h="left", align_v="center", wrap=False,
          italic=False, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ft(bold=bold, size=size, color=color, italic=italic)
    c.alignment = al(align_h, align_v, wrap)
    if bg:
        c.fill = fill(bg)
    if border is not None:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c

def header_row(ws, row, text, cols_merge_end=None, size=14):
    """Fila de título de sección."""
    c = write(ws, row, 1, text, bold=True, size=size,
              color=C_WHITE, bg=C_HEADER_SEC, align_h="center")
    if cols_merge_end:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols_merge_end)

def sub_header(ws, row, cols, labels):
    """Fila de cabecera de tabla."""
    for i, lbl in enumerate(labels):
        col = cols[i] if i < len(cols) else cols[-1] + i
        write(ws, row, col, lbl, bold=True, size=10,
              color=C_WHITE, bg=C_HEADER_TAB, align_h="center",
              border=border_thin())

def add_nav_button_note(ws, row, text="← Usa los botones de la hoja INICIO para navegar"):
    write(ws, row, 1, text, italic=True, size=9, color=C_MUTED)

# ── Datos de defecto del configurador ────────────────────────
ESPACIOS_DEFAULT = [
    ("SI", "bano",       "🛁 Baño",           "🛁", 6,  3, 20, 3),
    ("SI", "aseo",       "🚿 Aseo",            "🚿", 4,  2, 10, 2),
    ("SI", "cocina",     "🍳 Cocina",           "🍳", 12, 6, 30, 1),
    ("SI", "habitacion", "🛏️ Habitación",       "🛏️", 12, 8, 25, 5),
    ("SI", "salon",      "🛋️ Salón",            "🛋️", 20,10, 60, 1),
    ("NO", "pasillo",    "🚪 Pasillo",          "🚪", 8,  3, 20, 1),
    ("NO", "toda_la_obra","🏗️ Toda la obra",   "🏗️", 80,30,300, 1),
    ("NO", "terraza",    "🏡 Terraza",          "🏡", 15, 5, 80, 2),
]

TRABAJOS_DEFAULT = [
    # (activo, espacio, clave, label, unidad, mo, mat, dias, nota)
    # BAÑO
    ("SI","bano","demolicion_completa","Demolición completa","m2",40,5,1.0,""),
    ("SI","bano","solo_revestimientos","Solo retirada de revestimientos","m2",18,3,0.5,""),
    ("SI","bano","mover_distribucion","Redistribución de tabiques","ud",600,200,2.0,""),
    ("SI","bano","renovar_fontaneria","Renovar fontanería completa","ud",700,450,2.5,""),
    ("SI","bano","renovar_electricidad","Instalación eléctrica","ud",350,200,1.5,""),
    ("SI","bano","suelo_radiante","Suelo radiante","m2",35,40,0.5,""),
    ("SI","bano","cambiar_baldosas","Cambiar baldosas (zona húmeda/ducha)","m2_pared",38,48,0.6,"Incluye enfoscado e impermeabilización"),
    ("SI","bano","cambiar_suelo","Cambiar suelo del baño","m2",28,28,0.4,"Incluye retirada y preparación"),
    ("SI","bano","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","bano","falso_techo","Techo / Falso techo","m2",25,18,0.3,""),
    ("SI","bano","instalar_ducha","Instalación de ducha/plato","ud",200,350,1.0,""),
    ("SI","bano","instalar_banera","Instalación de bañera","ud",250,450,1.0,""),
    ("SI","bano","instalar_inodoro","Instalar inodoro","ud",120,200,0.5,""),
    ("SI","bano","instalar_lavabo","Instalar lavabo","ud",150,280,0.5,""),
    ("SI","bano","instalar_mueble","Montaje de mueble de baño","ud",100,200,0.5,""),
    ("SI","bano","instalar_espejo","Espejo o botiquín","ud",40,90,0.2,""),
    ("SI","bano","radiador_toallero","Radiador toallero","ud",80,140,0.3,""),
    ("SI","bano","cambiar_puerta","Cambiar puerta","ud",150,290,0.5,""),
    ("SI","bano","cambiar_ventana","Cambiar ventana","ud",200,430,0.5,""),
    # ASEO
    ("SI","aseo","demolicion_completa","Demolición completa","m2",40,5,0.5,""),
    ("SI","aseo","solo_revestimientos","Solo retirada de revestimientos","m2",18,3,0.3,""),
    ("SI","aseo","mover_distribucion","Redistribución de tabiques","ud",600,200,2.0,""),
    ("SI","aseo","renovar_fontaneria","Renovar fontanería","ud",500,350,2.0,""),
    ("SI","aseo","renovar_electricidad","Instalación eléctrica","ud",280,160,1.0,""),
    ("SI","aseo","cambiar_baldosas","Cambiar baldosas (zona húmeda)","m2_pared",38,48,0.6,"Incluye enfoscado e impermeabilización"),
    ("SI","aseo","cambiar_suelo","Cambiar suelo del aseo","m2",28,28,0.4,""),
    ("SI","aseo","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","aseo","falso_techo","Techo / Falso techo","m2",25,18,0.2,""),
    ("SI","aseo","instalar_ducha","Instalación de ducha","ud",200,350,1.0,""),
    ("SI","aseo","instalar_inodoro","Instalar inodoro","ud",120,200,0.5,""),
    ("SI","aseo","instalar_lavabo","Instalar lavabo","ud",150,230,0.5,""),
    ("SI","aseo","instalar_mueble","Montaje de mueble de baño/aseo","ud",100,200,0.5,""),
    ("SI","aseo","radiadores","Instalar radiador","ud",150,280,0.5,""),
    ("SI","aseo","cambiar_puerta","Cambiar puerta","ud",150,290,0.5,""),
    ("SI","aseo","cambiar_ventana","Cambiar ventana","ud",200,380,0.5,""),
    # COCINA
    ("SI","cocina","demolicion_completa","Demolición y vaciado","m2",45,5,1.0,""),
    ("SI","cocina","mover_distribucion","Redistribución de tabiques","ud",700,250,2.0,""),
    ("SI","cocina","renovar_fontaneria","Renovar fontanería","ud",600,400,2.0,""),
    ("SI","cocina","renovar_electricidad","Instalación eléctrica","ud",500,300,1.5,""),
    ("SI","cocina","alicatar_paredes","Cambiar azulejos / revestimiento","m2_pared",22,25,0.3,""),
    ("SI","cocina","cambiar_suelo","Cambiar suelo","m2",18,22,0.3,""),
    ("SI","cocina","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","cocina","falso_techo","Techo / Falso techo","m2",25,18,0.3,""),
    ("SI","cocina","instalar_muebles","Muebles de cocina","ml",180,350,1.5,""),
    ("SI","cocina","encimera","Encimera","ml",120,250,0.5,""),
    ("SI","cocina","instalar_campana","Campana extractora","ud",120,280,0.5,""),
    ("SI","cocina","instalar_fregadero","Fregadero","ud",150,220,0.5,""),
    ("SI","cocina","instalar_caldera","Instalar caldera","ud",350,800,1.5,""),
    ("SI","cocina","radiadores","Instalar radiador","ud",150,280,0.5,""),
    ("SI","cocina","cambiar_puerta","Cambiar puerta","ud",150,290,0.5,""),
    ("SI","cocina","cambiar_ventana","Cambiar ventana","ud",200,430,0.5,""),
    # HABITACION
    ("SI","habitacion","tirar_pared","Demolición de pared","ud",600,100,1.0,""),
    ("SI","habitacion","redistribucion_tabiques","Redistribución de tabiques","ud",900,350,2.5,""),
    ("SI","habitacion","abrir_hueco","Abrir o tapiar un hueco","ud",350,50,0.5,""),
    ("SI","habitacion","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","habitacion","papel_pintado","Papel pintado","m2_pared",12,18,0.2,""),
    ("SI","habitacion","cambiar_suelo","Cambiar suelo","m2",18,22,0.3,""),
    ("SI","habitacion","pulir_parquet","Pulir y barnizar parquet","m2",12,4,0.2,""),
    ("SI","habitacion","falso_techo","Falso techo o molduras","m2",25,18,0.3,""),
    ("SI","habitacion","cambiar_ventanas","Cambiar ventana(s)","ud",200,430,0.5,""),
    ("SI","habitacion","armario_empotrado","Armario empotrado","ml",200,380,1.0,""),
    ("SI","habitacion","cambiar_puerta","Cambiar puerta","ud",150,290,0.5,""),
    ("SI","habitacion","puntos_de_luz","Puntos de luz y enchufes","ud",80,30,0.3,""),
    ("SI","habitacion","aire_acondicionado","Aire acondicionado split","ud",300,900,1.0,""),
    ("SI","habitacion","suelo_radiante","Suelo radiante","m2",35,40,0.4,""),
    ("SI","habitacion","radiadores","Instalar radiador","ud",150,280,0.5,""),
    # SALON
    ("SI","salon","tirar_pared","Demolición de pared","ud",600,100,1.0,""),
    ("SI","salon","redistribucion_tabiques","Redistribución de tabiques","ud",900,350,2.5,""),
    ("SI","salon","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","salon","papel_pintado","Papel pintado","m2_pared",12,18,0.2,""),
    ("SI","salon","cambiar_suelo","Cambiar suelo","m2",18,22,0.3,""),
    ("SI","salon","pulir_parquet","Pulir y barnizar parquet","m2",12,4,0.2,""),
    ("SI","salon","falso_techo","Falso techo o molduras","m2",25,18,0.3,""),
    ("SI","salon","puntos_de_luz","Puntos de luz y enchufes","ud",80,30,0.3,""),
    ("SI","salon","aire_acondicionado","Aire acondicionado split","ud",300,900,1.0,""),
    ("SI","salon","cambiar_ventanas","Cambiar ventana(s)","ud",200,430,0.5,""),
    ("SI","salon","radiadores","Radiadores","ud",150,280,0.5,""),
    # PASILLO
    ("SI","pasillo","tirar_pared","Demolición de pared","ud",500,80,0.8,""),
    ("SI","pasillo","redistribucion_tabiques","Redistribución de tabiques","ud",800,300,2.0,""),
    ("SI","pasillo","pintar_paredes","Pintura de paredes y techo","m2_pared",8,4,0.2,""),
    ("SI","pasillo","cambiar_suelo","Cambiar suelo","m2",18,22,0.2,""),
    ("SI","pasillo","falso_techo","Falso techo","m2",25,18,0.2,""),
    ("SI","pasillo","cambiar_puerta","Cambiar puerta(s)","ud",150,290,0.5,""),
    ("SI","pasillo","radiadores","Instalar radiador","ud",150,280,0.5,""),
    # TODA LA OBRA
    ("SI","toda_la_obra","pintura_integral","Pintura integral vivienda","m2",7,4,0.1,""),
    ("SI","toda_la_obra","suelos_integral","Cambio de suelos completo","m2",18,22,0.2,""),
    ("SI","toda_la_obra","electricidad_integral","Instalación eléctrica compl.","m2",25,15,0.2,""),
    ("SI","toda_la_obra","fontaneria_integral","Fontanería completa","m2",20,12,0.2,""),
    ("SI","toda_la_obra","demolicion_integral","Demolición y redistribución","m2",35,5,0.2,""),
    # TERRAZA
    ("SI","terraza","limpiar_rejuntar","Limpiar y rejuntar baldosas","m2",8,3,0.1,""),
    ("SI","terraza","cambiar_suelo","Cambiar baldosas / suelo de terraza","m2",22,28,0.3,""),
    ("SI","terraza","impermeabilizar","Impermeabilización de terraza","m2",22,18,0.3,""),
    ("SI","terraza","pintar_paredes","Pintura de paredes / barandillas","m2_pared",8,4,0.2,""),
    ("SI","terraza","cerrar_terraza","Cerramiento de terraza","m2",85,130,1.0,""),
    ("SI","terraza","pergola_toldo","Pérgola o toldo","ud",200,400,1.0,""),
    ("SI","terraza","cambiar_barandilla","Cambiar barandilla","ml",80,120,0.5,""),
    ("SI","terraza","falso_techo","Techo / Falso techo","m2",25,18,0.3,""),
    ("SI","terraza","radiadores","Instalar radiador/calefactor","ud",150,280,0.5,""),
]

MATERIALES_DEFAULT = [
    # (activo, id, cat, nombre, marca, precio, unidad, proveedor, visible)
    ("SI","p1","suelo","Azulejo porcelánico Travertino Beige 60×60","Aparici",18.90,"m²","Obramart","SI"),
    ("SI","p2","suelo","Parquet laminado Roble Natural AC5 8mm","Kronotex",16.50,"m²","Obramart","SI"),
    ("SI","p3","suelo","Gres porcelánico Cemento Gris 80×80","Vives",24.30,"m²","Obramart","SI"),
    ("SI","p4","pintura","Pintura plástica blanca mate 15L","Titanlux",32.50,"l","Obramart","SI"),
    ("SI","p5","pintura","Esmalte satinado blanco paredes 4L","Valentine",18.90,"l","Obramart","SI"),
    ("SI","p6","tabiqueria","Placa Pladur N 15mm 120×260 cm","Pladur",11.20,"m²","Obramart","SI"),
    ("SI","p7","tabiqueria","Placa Pladur WA antihumedad 15mm","Pladur",14.80,"m²","Obramart","SI"),
    ("SI","p8","sanitario","Inodoro suspendido Rimless doble descarga","Roca",189.00,"ud","Leroy Merlin","SI"),
    ("SI","p9","sanitario","Lavabo sobre encimera 60cm ovalado blanco","Villeroy",145.00,"ud","Leroy Merlin","SI"),
    ("SI","p10","mueble","Mueble lavabo 80cm suspendido con cajones","Camargue",385.00,"ud","Leroy Merlin","SI"),
    ("SI","p11","fontaneria","Grifo monomando baño cromado caudal regulable","Grohe",89.00,"ud","Leroy Merlin","SI"),
    ("SI","p12","caldera","Caldera de condensación gas natural 24kW","Baxi",1290.00,"ud","Bricomart","SI"),
    ("SI","p13","calefaccion","Radiador acero tubular 600mm 6 elementos","Zehnder",145.00,"ud","Bricomart","SI"),
    ("SI","p14","suelo","Tarima flotante AC4 8mm Roble Natural","Quickstep",22.50,"m²","Leroy Merlin","SI"),
    ("SI","az1","azulejo","Azulejo metro blanco mate 10×20","Equipe",22.00,"m²","Obramart","SI"),
    ("SI","az2","azulejo","Azulejo hidráulico azul 20×20","Realonda",28.50,"m²","Obramart","SI"),
]

EXTRAS_DEFAULT = [
    # (activo, clave, label, desc, icon, importe, dinamico)
    ("SI","sin_ascensor","Edificio sin ascensor","Se aplica recargo por planta según el piso seleccionado (3–9% s/total)","🏗️",0,"SI"),
    ("SI","gestion_residuos","Gestión de residuos y escombros","Retirada y vertedero incluidos","🗑️",280,"NO"),
    ("SI","necesita_grua","Necesito grúa o montacargas","Para muebles o maquinaria pesada","🏋️",450,"NO"),
    ("SI","proteccion_zonas","Protección de zonas no reformadas","Mobiliario y suelos protegidos","🛡️",180,"NO"),
    ("SI","obra_habitada","Obra en vivienda habitada","Requiere coordinación especial","🏠",350,"NO"),
    ("SI","autorizacion_comunidad","Gestión autorización de comunidad","Trámite con el administrador","📋",80,"NO"),
    ("SI","sin_parking","Sin parking para furgoneta","Desplazamiento desde parking lejano","🅿️",120,"NO"),
]

# ── Colores de fila alternada ─────────────────────────────────
def row_fill(i):
    return C_WHITE if i % 2 == 0 else C_LIGHT

# ─────────────────────────────────────────────────────────────
# HOJA: INICIO (portada + navegación)
# ─────────────────────────────────────────────────────────────
def build_inicio(wb):
    ws = wb.create_sheet("INICIO")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 4

    # Título principal
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 60
    ws.merge_cells("B2:D2")
    c = write(ws, 2, 2, "ONBOARDING REFORMISTA", bold=True, size=26,
              color=C_WHITE, bg=C_BRAND, align_h="center")
    ws.merge_cells("B3:D3")
    write(ws, 3, 2, "Guía de configuración del presupuestador", italic=True, size=13,
          color=C_ACCENT, bg=C_BRAND, align_h="center")
    ws.row_dimensions[3].height = 30

    # Descripción
    ws.row_dimensions[5].height = 8
    ws.merge_cells("B6:D7")
    write(ws, 6, 2, (
        "Completa cada hoja con los datos del reformista.\n"
        "Al terminar, haz clic en 'EXPORTAR JSON' para generar el archivo\n"
        "que se importará en el Panel de Administración."
    ), size=11, color=C_LABEL, bg=C_WHITE, align_h="center", wrap=True)
    ws.row_dimensions[6].height = 55

    # Sección NAVEGACIÓN
    ws.row_dimensions[9].height = 8
    ws.merge_cells("B10:D10")
    write(ws, 10, 2, "SECCIONES DEL FORMULARIO", bold=True, size=13,
          color=C_WHITE, bg=C_HEADER_SEC, align_h="center")
    ws.row_dimensions[10].height = 30

    secciones = [
        ("🏢 EMPRESA",      "Datos de la empresa y acceso al panel"),
        ("⚙️ MOTOR",        "Márgenes, IVA, mínimos y calidades"),
        ("🏠 ESPACIOS",     "Tipos de espacio disponibles"),
        ("🔨 TRABAJOS",     "Tipos de trabajo por espacio y precios"),
        ("🧱 MATERIALES",   "Catálogo de productos y materiales"),
        ("➕ EXTRAS",       "Servicios adicionales y cargos"),
        ("📣 COMUNICACION", "Mensajes, WhatsApp y marca"),
        ("📍 ZONA",         "Cobertura geográfica y desplazamiento"),
        ("📤 EXPORTAR",     "Generar JSON e importar en el panel"),
    ]

    for i, (titulo, desc) in enumerate(secciones):
        r = 11 + i
        ws.row_dimensions[r].height = 26
        bg = row_fill(i)
        write(ws, r, 2, titulo, bold=True, size=11, color=C_HEADER_TAB, bg=bg,
              border=border_thin())
        write(ws, r, 3, desc, size=10, color=C_LABEL, bg=bg, border=border_thin())

        # Nota: Los botones de navegación reales son macros VBA.
        # En las celdas D ponemos el nombre de la hoja como referencia.
        sheet_name = titulo.split(" ", 1)[1].strip() if " " in titulo else titulo
        write(ws, r, 4, f"→ Ir a {sheet_name}", size=10, color=C_ACCENT, bg=bg,
              align_h="center", border=border_thin())

    # Pie
    r_pie = 11 + len(secciones) + 2
    ws.merge_cells(f"B{r_pie}:D{r_pie}")
    write(ws, r_pie, 2, (
        "💡 Tip: Habilita las macros al abrir este archivo. "
        "Las macros generan el JSON de configuración."
    ), italic=True, size=9, color=C_MUTED, align_h="center")

# ─────────────────────────────────────────────────────────────
# HOJA: EMPRESA
# ─────────────────────────────────────────────────────────────
def build_empresa(wb):
    ws = wb.create_sheet("EMPRESA")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 20

    header_row(ws, 1, "🏢  DATOS DE LA EMPRESA", 4)
    ws.row_dimensions[1].height = 35

    write(ws, 2, 1, "Rellena los datos de la empresa del reformista.",
          italic=True, size=10, color=C_MUTED)

    # Bloque: Datos empresa
    ws.row_dimensions[3].height = 22
    write(ws, 3, 2, "DATOS EMPRESARIALES", bold=True, size=10,
          color=C_WHITE, bg=C_HEADER_TAB)

    campos_empresa = [
        ("Nombre de la empresa *",       "Anatoliy Reformas S.L.",    "B4:B4"),
        ("CIF / NIF",                    "B12345678",                 ""),
        ("Teléfono de contacto",         "+34 612 345 678",           ""),
        ("Email de contacto *",          "info@anatoliyreformas.es",  ""),
        ("Página web",                   "www.anatoliyreformas.es",   ""),
        ("Ciudad principal",             "Madrid",                    ""),
        ("Dirección",                    "Calle Mayor 12, 28001 Madrid",""),
        ("Nombre del autónomo",          "Anatoliy Marchenko",        ""),
    ]
    for i, (lbl, ejemplo, _) in enumerate(campos_empresa):
        r = 4 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        write(ws, r, 2, lbl, bold=("*" in lbl), size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 3, "", size=11, color=C_LABEL, bg=C_WHITE, border=border_thin())
        write(ws, r, 4, f"Ej: {ejemplo}", italic=True, size=9, color=C_MUTED, bg=bg)

    # Bloque: Acceso panel admin
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 22
    write(ws, 13, 2, "ACCESO AL PANEL DE ADMINISTRACIÓN", bold=True, size=10,
          color=C_WHITE, bg=C_ACCENT)

    campos_acceso = [
        ("Email de acceso al panel *",   "admin@anatoliyreformas.es"),
        ("Contraseña de acceso *",       "MiClaveSegura2024!"),
    ]
    for i, (lbl, ejemplo) in enumerate(campos_acceso):
        r = 14 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        write(ws, r, 2, lbl, bold=True, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 3, "", size=11, color=C_LABEL, bg=C_WHITE, border=border_thin())
        write(ws, r, 4, f"Ej: {ejemplo}", italic=True, size=9, color=C_MUTED, bg=bg)

    write(ws, 16, 2, "⚠️  La contraseña se almacena en texto plano en este Excel. Trátalo como documento confidencial.",
          italic=True, size=9, color=C_YELLOW)

    add_nav_button_note(ws, 18)

# ─────────────────────────────────────────────────────────────
# HOJA: MOTOR
# ─────────────────────────────────────────────────────────────
def build_motor(wb):
    ws = wb.create_sheet("MOTOR")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    header_row(ws, 1, "⚙️  MOTOR DE CÁLCULO", 4)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1, "Parámetros que determinan cómo se calcula cada presupuesto.",
          italic=True, size=10, color=C_MUTED)

    # Sección: Margen y fiscalidad
    ws.row_dimensions[3].height = 22
    write(ws, 3, 2, "MÁRGENES Y FISCALIDAD", bold=True, size=10,
          color=C_WHITE, bg=C_HEADER_TAB)

    campos_motor = [
        # (label, valor_defecto, formato, ayuda)
        ("Tipo de margen", "porcentaje_total",
         None, "porcentaje_total · margen_fijo · por_partida"),
        ("Margen (%)", 22,
         "0.0", "Porcentaje sobre el total (ej: 22 = 22%). Introducir como número, no como porcentaje."),
        ("IVA (%)", 21,
         "0", "IVA aplicado en la factura (España: 10% reformas, 21% general). Número entero."),
        ("Presupuesto mínimo (€)", 500,
         "#,##0 €", "El presupuesto no puede salir por debajo de este importe"),
        ("Validez del presupuesto (días)", 30,
         "0", "Días que es válido el presupuesto desde su emisión"),
        ("Desplazamiento base (€)", 80,
         "#,##0 €", "Coste fijo de desplazamiento incluido en todos los presupuestos"),
        ("Garantía mano de obra (meses)", 12,
         "0", "Meses de garantía sobre los trabajos realizados"),
        ("Garantía materiales (meses)", 24,
         "0", "Meses de garantía sobre los materiales instalados"),
    ]

    for i, (lbl, val, fmt, ayuda) in enumerate(campos_motor):
        r = 4 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        write(ws, r, 2, lbl, bold=False, size=10, color=C_LABEL, bg=bg, border=border_thin())
        c = ws.cell(row=r, column=3, value=val)
        c.font = ft(bold=True, size=11, color=C_HEADER_TAB)
        c.fill = fill(C_WHITE)
        c.alignment = al("center")
        c.border = border_thin()
        if fmt:
            c.number_format = fmt
        write(ws, r, 4, ayuda, italic=True, size=9, color=C_MUTED, bg=bg)

    # Validación: margen_tipo con lista desplegable
    dv_margen = DataValidation(
        type="list",
        formula1='"porcentaje_total,margen_fijo,por_partida"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Tipo no válido",
        error='Elige: porcentaje_total, margen_fijo o por_partida'
    )
    ws.add_data_validation(dv_margen)
    dv_margen.add("C4")

    # Sección: Calidades
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 22
    write(ws, 13, 2, "COEFICIENTES DE CALIDAD", bold=True, size=10,
          color=C_WHITE, bg=C_HEADER_TAB)
    write(ws, 13, 3, "Coeficiente", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB, align_h="center")
    write(ws, 13, 4, "Efecto sobre precio base", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)

    calidades = [
        ("Básica (coef.)", 0.80, "Reduce el precio en un 20% sobre la base estándar"),
        ("Media / Estándar (coef.)", 1.00, "Precio base de referencia — sin modificación"),
        ("Premium (coef.)", 1.40, "Incrementa el precio un 40% sobre la base estándar"),
    ]
    for i, (lbl, val, ayuda) in enumerate(calidades):
        r = 14 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        write(ws, r, 2, lbl, size=10, color=C_LABEL, bg=bg, border=border_thin())
        c = ws.cell(row=r, column=3, value=val)
        c.font = ft(bold=True, size=11, color=C_GREEN)
        c.fill = fill(C_WHITE)
        c.alignment = al("center")
        c.border = border_thin()
        c.number_format = "0.00"
        write(ws, r, 4, ayuda, italic=True, size=9, color=C_MUTED, bg=bg)

    # Sección: Forma de pago
    ws.row_dimensions[17].height = 18
    ws.row_dimensions[18].height = 22
    write(ws, 18, 2, "CONDICIONES DE PAGO", bold=True, size=10,
          color=C_WHITE, bg=C_HEADER_TAB)

    campos_pago = [
        ("Forma de pago", "50_50", None, "50_50 · 30_70 · 100_inicio · 100_fin · personalizado"),
        ("Días de pago (plazo)", 30, "0", "Días desde la emisión para el pago final"),
        ("Penalización por retraso (%)", 1.5, "0.0", "Interés mensual por pago tardío (ej: 1.5 = 1.5%). Número decimal."),
    ]
    for i, (lbl, val, fmt, ayuda) in enumerate(campos_pago):
        r = 19 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        write(ws, r, 2, lbl, size=10, color=C_LABEL, bg=bg, border=border_thin())
        c = ws.cell(row=r, column=3, value=val)
        c.font = ft(bold=True, size=11, color=C_HEADER_TAB)
        c.fill = fill(C_WHITE)
        c.alignment = al("center")
        c.border = border_thin()
        if fmt:
            c.number_format = fmt
        write(ws, r, 4, ayuda, italic=True, size=9, color=C_MUTED, bg=bg)

    # Validación forma de pago
    dv_pago = DataValidation(
        type="list",
        formula1='"50_50,30_70,100_inicio,100_fin,personalizado"',
        allow_blank=False
    )
    ws.add_data_validation(dv_pago)
    dv_pago.add("C19")

    add_nav_button_note(ws, 23)

# ─────────────────────────────────────────────────────────────
# HOJA: ESPACIOS
# ─────────────────────────────────────────────────────────────
def build_espacios(wb):
    ws = wb.create_sheet("ESPACIOS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8   # activo
    ws.column_dimensions['C'].width = 16  # id
    ws.column_dimensions['D'].width = 22  # label
    ws.column_dimensions['E'].width = 8   # emoji
    ws.column_dimensions['F'].width = 12  # m2_defecto
    ws.column_dimensions['G'].width = 10  # m2_min
    ws.column_dimensions['H'].width = 10  # m2_max
    ws.column_dimensions['I'].width = 14  # max_instancias
    ws.column_dimensions['J'].width = 30  # notas

    header_row(ws, 1, "🏠  ESPACIOS DISPONIBLES", 10)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1,
          "Marca SI/NO para activar cada espacio. El cliente solo verá los espacios activos.",
          italic=True, size=10, color=C_MUTED)

    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 28
    sub_header(ws, 4, list(range(2, 11)),
               ["Activo", "ID espacio", "Etiqueta", "Emoji",
                "m² defecto", "m² mín.", "m² máx.", "Máx. instancias", "Notas"])

    # Validación SI/NO
    dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv_sino)

    for i, (act, eid, label, emoji, m2d, m2min, m2max, max_inst) in enumerate(ESPACIOS_DEFAULT):
        r = 5 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)
        color_act = C_GREEN if act == "SI" else C_MUTED

        c_act = ws.cell(row=r, column=2, value=act)
        c_act.font = ft(bold=True, size=10, color=color_act)
        c_act.fill = fill(bg)
        c_act.alignment = al("center")
        c_act.border = border_thin()
        dv_sino.add(f"B{r}")

        write(ws, r, 3, eid, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 4, label, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 5, emoji, size=14, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 6, m2d, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 7, m2min, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 8, m2max, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 9, max_inst, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 10, "", size=10, color=C_MUTED, bg=bg, border=border_thin())

    add_nav_button_note(ws, 5 + len(ESPACIOS_DEFAULT) + 2)

# ─────────────────────────────────────────────────────────────
# HOJA: TRABAJOS
# ─────────────────────────────────────────────────────────────
def build_trabajos(wb):
    ws = wb.create_sheet("TRABAJOS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 7   # activo
    ws.column_dimensions['B'].width = 2
    ws.column_dimensions['C'].width = 14  # espacio
    ws.column_dimensions['D'].width = 28  # clave
    ws.column_dimensions['E'].width = 38  # label
    ws.column_dimensions['F'].width = 12  # unidad
    ws.column_dimensions['G'].width = 10  # mo
    ws.column_dimensions['H'].width = 10  # mat
    ws.column_dimensions['I'].width = 10  # dias
    ws.column_dimensions['J'].width = 40  # nota

    header_row(ws, 1, "🔨  TIPOS DE TRABAJO Y PRECIOS", 10)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1,
          "Un trabajo = una fila. MO = mano de obra €/unidad. MAT = materiales €/unidad. DÍAS = tiempo estimado/unidad.",
          italic=True, size=10, color=C_MUTED)
    write(ws, 3, 1,
          "Unidades: m2 (área suelo), m2_pared (paredes+techo estimadas), ud (unidades), ml (metros lineales)",
          italic=True, size=9, color=C_MUTED)

    ws.row_dimensions[4].height = 28
    sub_header(ws, 4, [1, 3, 4, 5, 6, 7, 8, 9, 10],
               ["Activo", "Espacio", "Clave trabajo", "Descripción",
                "Unidad", "MO (€)", "MAT (€)", "Días/ud", "Nota"])

    dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv_sino)

    espacios_order = ["bano","aseo","cocina","habitacion","salon","pasillo","toda_la_obra","terraza"]
    cur_esp = None
    row_i = 0

    for (act, esp, clave, label, unidad, mo, mat, dias, nota) in TRABAJOS_DEFAULT:
        r = 5 + row_i
        ws.row_dimensions[r].height = 22
        bg = row_fill(row_i)

        if esp != cur_esp:
            # Separador de espacio
            if cur_esp is not None:
                ws.row_dimensions[r].height = 8
                row_i += 1
                r = 5 + row_i
                ws.row_dimensions[r].height = 22

            # Cabecera de grupo de espacio
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            lbl_esp = {
                "bano":"🛁 BAÑO","aseo":"🚿 ASEO","cocina":"🍳 COCINA",
                "habitacion":"🛏️ HABITACIÓN","salon":"🛋️ SALÓN",
                "pasillo":"🚪 PASILLO","toda_la_obra":"🏗️ TODA LA OBRA",
                "terraza":"🏡 TERRAZA"
            }.get(esp, esp.upper())
            write(ws, r, 1, lbl_esp, bold=True, size=10,
                  color=C_WHITE, bg=C_HEADER_SEC, align_h="center")
            cur_esp = esp
            row_i += 1
            r = 5 + row_i
            ws.row_dimensions[r].height = 22
            bg = row_fill(row_i)

        color_act = C_GREEN if act == "SI" else C_MUTED
        c_act = ws.cell(row=r, column=1, value=act)
        c_act.font = ft(bold=True, size=10, color=color_act)
        c_act.fill = fill(bg)
        c_act.alignment = al("center")
        c_act.border = border_thin()
        dv_sino.add(f"A{r}")

        write(ws, r, 3, esp, size=9, color=C_MUTED, bg=bg, border=border_thin())
        write(ws, r, 4, clave, size=9, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 5, label, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 6, unidad, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())

        c_mo = ws.cell(row=r, column=7, value=mo)
        c_mo.font = ft(size=10, color=C_GREEN)
        c_mo.fill = fill(C_WHITE)
        c_mo.alignment = al("right")
        c_mo.border = border_thin()
        c_mo.number_format = "#,##0 €"

        c_mat = ws.cell(row=r, column=8, value=mat)
        c_mat.font = ft(size=10, color=C_ACCENT)
        c_mat.fill = fill(C_WHITE)
        c_mat.alignment = al("right")
        c_mat.border = border_thin()
        c_mat.number_format = "#,##0 €"

        c_dias = ws.cell(row=r, column=9, value=dias)
        c_dias.font = ft(size=10, color=C_LABEL)
        c_dias.fill = fill(bg)
        c_dias.alignment = al("center")
        c_dias.border = border_thin()
        c_dias.number_format = "0.0"

        write(ws, r, 10, nota, italic=True, size=9, color=C_MUTED, bg=bg, border=border_thin())

        row_i += 1

    add_nav_button_note(ws, 5 + row_i + 2)

# ─────────────────────────────────────────────────────────────
# HOJA: MATERIALES
# ─────────────────────────────────────────────────────────────
def build_materiales(wb):
    ws = wb.create_sheet("MATERIALES")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 7   # activo
    ws.column_dimensions['C'].width = 8   # id
    ws.column_dimensions['D'].width = 14  # cat
    ws.column_dimensions['E'].width = 40  # nombre
    ws.column_dimensions['F'].width = 14  # marca
    ws.column_dimensions['G'].width = 10  # precio
    ws.column_dimensions['H'].width = 8   # unidad
    ws.column_dimensions['I'].width = 16  # proveedor
    ws.column_dimensions['J'].width = 8   # visible

    header_row(ws, 1, "🧱  CATÁLOGO DE MATERIALES Y PRODUCTOS", 10)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1,
          "Lista de materiales disponibles. El cliente los verá al configurar cada espacio. Precio = coste de compra (sin margen).",
          italic=True, size=10, color=C_MUTED)
    write(ws, 3, 1,
          "Categorías típicas: suelo · azulejo · pintura · tabiqueria · sanitario · mueble · fontaneria · caldera · calefaccion",
          italic=True, size=9, color=C_MUTED)

    ws.row_dimensions[4].height = 28
    sub_header(ws, 4, list(range(2, 11)),
               ["Activo", "ID", "Categoría", "Nombre del producto",
                "Marca", "Precio coste", "Unidad", "Proveedor", "Visible cliente"])

    dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv_sino)

    for i, (act, pid, cat, nombre, marca, precio, unidad, proveedor, visible) in enumerate(MATERIALES_DEFAULT):
        r = 5 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)

        color_act = C_GREEN if act == "SI" else C_MUTED
        c_act = ws.cell(row=r, column=2, value=act)
        c_act.font = ft(bold=True, size=10, color=color_act)
        c_act.fill = fill(bg)
        c_act.alignment = al("center")
        c_act.border = border_thin()
        dv_sino.add(f"B{r}")

        write(ws, r, 3, pid, size=9, color=C_MUTED, bg=bg, border=border_thin())
        write(ws, r, 4, cat, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 5, nombre, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 6, marca, size=10, color=C_MUTED, bg=bg, border=border_thin())

        c_precio = ws.cell(row=r, column=7, value=precio)
        c_precio.font = ft(bold=True, size=10, color=C_GREEN)
        c_precio.fill = fill(C_WHITE)
        c_precio.alignment = al("right")
        c_precio.border = border_thin()
        c_precio.number_format = "#,##0.00 €"

        write(ws, r, 8, unidad, size=10, color=C_LABEL, bg=bg, align_h="center", border=border_thin())
        write(ws, r, 9, proveedor, size=10, color=C_LABEL, bg=bg, border=border_thin())

        color_vis = C_GREEN if visible == "SI" else C_MUTED
        c_vis = ws.cell(row=r, column=10, value=visible)
        c_vis.font = ft(bold=True, size=10, color=color_vis)
        c_vis.fill = fill(bg)
        c_vis.alignment = al("center")
        c_vis.border = border_thin()
        dv_sino.add(f"J{r}")

    add_nav_button_note(ws, 5 + len(MATERIALES_DEFAULT) + 2)

# ─────────────────────────────────────────────────────────────
# HOJA: EXTRAS
# ─────────────────────────────────────────────────────────────
def build_extras(wb):
    ws = wb.create_sheet("EXTRAS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 7   # activo
    ws.column_dimensions['C'].width = 24  # clave
    ws.column_dimensions['D'].width = 30  # label
    ws.column_dimensions['E'].width = 45  # desc
    ws.column_dimensions['F'].width = 7   # icon
    ws.column_dimensions['G'].width = 12  # importe
    ws.column_dimensions['H'].width = 10  # dinamico

    header_row(ws, 1, "➕  SERVICIOS EXTRA Y CARGOS ADICIONALES", 8)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1,
          "Extras que el cliente puede seleccionar al configurar su presupuesto. Dinámico = el importe varía según el presupuesto.",
          italic=True, size=10, color=C_MUTED)

    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 28
    sub_header(ws, 4, list(range(2, 9)),
               ["Activo", "Clave (ID)", "Etiqueta visible", "Descripción",
                "Icono", "Importe (€)", "Dinámico"])

    dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv_sino)

    for i, (act, clave, label, desc, icon, importe, dinamico) in enumerate(EXTRAS_DEFAULT):
        r = 5 + i
        ws.row_dimensions[r].height = 24
        bg = row_fill(i)

        color_act = C_GREEN if act == "SI" else C_MUTED
        c_act = ws.cell(row=r, column=2, value=act)
        c_act.font = ft(bold=True, size=10, color=color_act)
        c_act.fill = fill(bg)
        c_act.alignment = al("center")
        c_act.border = border_thin()
        dv_sino.add(f"B{r}")

        write(ws, r, 3, clave, size=9, color=C_MUTED, bg=bg, border=border_thin())
        write(ws, r, 4, label, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 5, desc, size=9, color=C_MUTED, bg=bg, wrap=True, border=border_thin())
        write(ws, r, 6, icon, size=14, color=C_LABEL, bg=bg, align_h="center", border=border_thin())

        c_imp = ws.cell(row=r, column=7, value=importe)
        c_imp.font = ft(bold=True, size=10, color=C_ACCENT)
        c_imp.fill = fill(C_WHITE)
        c_imp.alignment = al("right")
        c_imp.border = border_thin()
        c_imp.number_format = "#,##0 €"

        color_din = C_YELLOW if dinamico == "SI" else C_MUTED
        c_din = ws.cell(row=r, column=8, value=dinamico)
        c_din.font = ft(bold=True, size=10, color=color_din)
        c_din.fill = fill(bg)
        c_din.alignment = al("center")
        c_din.border = border_thin()
        dv_sino.add(f"H{r}")

    add_nav_button_note(ws, 5 + len(EXTRAS_DEFAULT) + 2)

# ─────────────────────────────────────────────────────────────
# HOJA: COMUNICACION
# ─────────────────────────────────────────────────────────────
def build_comunicacion(wb):
    ws = wb.create_sheet("COMUNICACION")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25

    header_row(ws, 1, "📣  COMUNICACIÓN Y MARCA", 4)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1, "Textos y datos de contacto que aparecen en el configurador.",
          italic=True, size=10, color=C_MUTED)

    campos = [
        ("Mensaje de bienvenida", "Solicita tu presupuesto online en 3 minutos",
         "Texto que ve el cliente en el primer paso del configurador"),
        ("Mensaje de envío del presupuesto", "Tu presupuesto orientativo está listo. En breve te contactamos.",
         "Texto que aparece tras enviar el formulario"),
        ("WhatsApp de contacto", "+34 612 345 678",
         "Número con prefijo internacional (sin espacios)"),
        ("Color primario (hex)", "#1A1A2E",
         "Color principal de la interfaz (código hexadecimal)"),
        ("URL del logo", "",
         "Enlace a la imagen del logo (https://...)"),
    ]

    write(ws, 3, 2, "CAMPO", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)
    write(ws, 3, 3, "VALOR", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)
    write(ws, 3, 4, "AYUDA", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)

    for i, (lbl, val, ayuda) in enumerate(campos):
        r = 4 + i
        ws.row_dimensions[r].height = 28
        bg = row_fill(i)
        write(ws, r, 2, lbl, size=10, color=C_LABEL, bg=bg, border=border_thin())
        write(ws, r, 3, val, size=11, color=C_LABEL, bg=C_WHITE, border=border_thin(), wrap=True)
        write(ws, r, 4, ayuda, italic=True, size=9, color=C_MUTED, bg=bg, wrap=True)

    add_nav_button_note(ws, 10)

# ─────────────────────────────────────────────────────────────
# HOJA: ZONA
# ─────────────────────────────────────────────────────────────
def build_zona(wb):
    ws = wb.create_sheet("ZONA")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 35

    header_row(ws, 1, "📍  ZONA DE COBERTURA", 4)
    ws.row_dimensions[1].height = 35
    write(ws, 2, 1, "Área geográfica donde opera el reformista y costes de desplazamiento.",
          italic=True, size=10, color=C_MUTED)

    campos = [
        ("Radio de cobertura (km)", 30, "#,##0", "Kilómetros desde la base del reformista"),
        ("Coste por km adicional (€)", 0.35, "#,##0.00 €", "Coste extra por km fuera del desplazamiento base"),
        ("Municipios cubiertos", "Madrid, Alcalá de Henares, Getafe, Leganés",
         None, "Lista separada por comas de los municipios de cobertura"),
    ]

    write(ws, 3, 2, "CAMPO", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)
    write(ws, 3, 3, "VALOR", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)
    write(ws, 3, 4, "AYUDA", bold=True, size=10, color=C_WHITE, bg=C_HEADER_TAB)

    for i, row_data in enumerate(campos):
        lbl, val, fmt, ayuda = row_data
        r = 4 + i
        ws.row_dimensions[r].height = 30
        bg = row_fill(i)
        write(ws, r, 2, lbl, size=10, color=C_LABEL, bg=bg, border=border_thin())
        c = ws.cell(row=r, column=3, value=val)
        c.font = ft(bold=True, size=11, color=C_LABEL)
        c.fill = fill(C_WHITE)
        c.alignment = al("center", wrap=True)
        c.border = border_thin()
        if fmt:
            c.number_format = fmt
        write(ws, r, 4, ayuda, italic=True, size=9, color=C_MUTED, bg=bg, wrap=True)

    add_nav_button_note(ws, 8)

# ─────────────────────────────────────────────────────────────
# HOJA: EXPORTAR
# ─────────────────────────────────────────────────────────────
def build_exportar(wb):
    ws = wb.create_sheet("EXPORTAR")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 4

    # Header
    ws.merge_cells("B1:D1")
    c = write(ws, 1, 2, "📤  EXPORTAR CONFIGURACIÓN", bold=True, size=20,
              color=C_WHITE, bg=C_BRAND, align_h="center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("B2:D2")
    write(ws, 2, 2,
          "Genera el archivo JSON listo para importar en el Panel de Administración",
          italic=True, size=11, color=C_MUTED, align_h="center")
    ws.row_dimensions[2].height = 30

    ws.row_dimensions[4].height = 10

    # Instrucciones
    instrucciones = [
        ("1️⃣", "Completa todas las hojas del formulario"),
        ("2️⃣", "Haz clic en el botón 'VALIDAR FORMULARIO' para comprobar que no faltan datos"),
        ("3️⃣", "Haz clic en 'GENERAR Y GUARDAR JSON' para exportar el archivo"),
        ("4️⃣", "Accede al Panel de Administración → Configuración → Importar JSON"),
        ("5️⃣", "Selecciona el archivo generado y confirma la importación"),
    ]

    for i, (num, texto) in enumerate(instrucciones):
        r = 5 + i
        ws.row_dimensions[r].height = 30
        bg = row_fill(i)
        write(ws, r, 2, num, bold=True, size=14, color=C_ACCENT,
              bg=bg, align_h="center", border=border_thin())
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        write(ws, r, 3, texto, size=11, color=C_LABEL, bg=bg, border=border_thin())

    ws.row_dimensions[11].height = 20

    # Botones (como celdas formateadas — las macros VBA los activan)
    ws.merge_cells("B12:C12")
    c_val = write(ws, 12, 2, "✅  VALIDAR FORMULARIO",
                  bold=True, size=14, color=C_WHITE, bg=C_GREEN,
                  align_h="center")
    ws.row_dimensions[12].height = 45

    ws.merge_cells("B14:C14")
    c_gen = write(ws, 14, 2, "📥  GENERAR Y GUARDAR JSON",
                  bold=True, size=14, color=C_WHITE, bg=C_ACCENT,
                  align_h="center")
    ws.row_dimensions[14].height = 45

    ws.merge_cells("B16:C16")
    c_prev = write(ws, 16, 2, "👁️  PREVISUALIZAR JSON",
                   bold=True, size=12, color=C_WHITE, bg=C_HEADER_TAB,
                   align_h="center")
    ws.row_dimensions[16].height = 35

    ws.merge_cells("B18:D18")
    write(ws, 18, 2,
          "⚠️  Nota: Los botones solo funcionan con macros habilitadas (archivo .xlsm).",
          italic=True, size=10, color=C_YELLOW, align_h="center")

    ws.merge_cells("B19:D19")
    write(ws, 19, 2,
          "Al abrir el archivo, Excel pedirá habilitar macros. Haz clic en 'Habilitar contenido'.",
          italic=True, size=9, color=C_MUTED, align_h="center")

    add_nav_button_note(ws, 21)

# ─────────────────────────────────────────────────────────────
# EMBED VBA macro en el xlsm
# ─────────────────────────────────────────────────────────────
VBA_MODULE_CONTENT = open(
    os.path.join(os.path.dirname(__file__), "macros_onboarding.bas"),
    encoding="utf-8"
).read()

def embed_vba(xlsm_path, bas_path):
    """
    Embebe el módulo VBA dentro del xlsm usando manipulación ZIP.
    Esta técnica requiere el archivo .bas externo que openpyxl
    no puede insertar directamente. Generamos un vbaProject.bin
    usando xlsxwriter como backend alternativo, o documentamos
    que la inserción manual es necesaria.

    Para entornos sin win32com, generamos el archivo .bas separado
    y documentamos el proceso de importación manual.
    """
    # Intentar embeber con xlwt/win32com si está disponible
    try:
        import win32com.client
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        wb = xl.Workbooks.Open(os.path.abspath(xlsm_path))
        vba_mod = wb.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
        vba_mod.Name = "MacrosOnboarding"
        vba_mod.CodeModule.AddFromString(VBA_MODULE_CONTENT)
        # Botones en EXPORTAR
        _add_buttons_via_com(xl, wb)
        wb.Save()
        wb.Close()
        xl.Quit()
        return True
    except Exception:
        return False

def _add_buttons_via_com(xl, wb):
    """Asigna macros a las celdas-botón en la hoja EXPORTAR."""
    try:
        ws_exp = wb.Sheets("EXPORTAR")
        # Botón Validar (fila 12, col B:C)
        btn_val = ws_exp.Buttons().Add(
            ws_exp.Cells(12, 2).Left, ws_exp.Cells(12, 2).Top,
            ws_exp.Range("B12:C12").Width, 45
        )
        btn_val.OnAction = "ValidarFormulario"
        btn_val.Caption = "✅  VALIDAR FORMULARIO"
        # Botón Generar (fila 14)
        btn_gen = ws_exp.Buttons().Add(
            ws_exp.Cells(14, 2).Left, ws_exp.Cells(14, 2).Top,
            ws_exp.Range("B14:C14").Width, 45
        )
        btn_gen.OnAction = "GuardarJSON"
        btn_gen.Caption = "📥  GENERAR Y GUARDAR JSON"
        # Botón Preview (fila 16)
        btn_prev = ws_exp.Buttons().Add(
            ws_exp.Cells(16, 2).Left, ws_exp.Cells(16, 2).Top,
            ws_exp.Range("B16:C16").Width, 35
        )
        btn_prev.OnAction = "PreviewJSON"
        btn_prev.Caption = "👁️  PREVISUALIZAR JSON"
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    out_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(out_dir, "onboarding_reformista.xlsx")
    xlsm_path = os.path.join(out_dir, "onboarding_reformista.xlsm")

    print("Creando estructura Excel...")
    wb = Workbook()

    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_inicio(wb)
    build_empresa(wb)
    build_motor(wb)
    build_espacios(wb)
    build_trabajos(wb)
    build_materiales(wb)
    build_extras(wb)
    build_comunicacion(wb)
    build_zona(wb)
    build_exportar(wb)

    # Guardar xlsx base
    wb.save(xlsx_path)
    print(f"  ✓ Excel base guardado: {xlsx_path}")

    # Intentar crear xlsm (Excel con macros)
    # openpyxl no soporta VBA nativo; copiamos el xlsx con extensión xlsm
    # El usuario debe importar el .bas manualmente en el VBE (Alt+F11)
    shutil.copy(xlsx_path, xlsm_path)
    print(f"  ✓ Archivo xlsm copiado: {xlsm_path}")

    # Intentar embeber VBA si win32com está disponible (Windows)
    vba_embedded = embed_vba(xlsm_path, os.path.join(out_dir, "macros_onboarding.bas"))

    print()
    if vba_embedded:
        print("✅ VBA embebido correctamente en el archivo xlsm.")
        print(f"   Archivo listo: {xlsm_path}")
    else:
        print("⚠️  VBA no pudo embeberse automáticamente (requiere Windows + Excel).")
        print()
        print("PASOS PARA ACTIVAR LAS MACROS MANUALMENTE:")
        print("  1. Abre 'onboarding_reformista.xlsm' en Excel (Windows)")
        print("  2. Presiona Alt + F11 para abrir el Editor de Visual Basic")
        print("  3. En el menú: Archivo → Importar archivo...")
        print(f"  4. Selecciona: macros_onboarding.bas")
        print("  5. Cierra el editor y guarda el archivo como .xlsm")
        print()
        print("  O desde Excel: Desarrollador → Visual Basic → Insertar → Módulo")
        print("  y pega el contenido de macros_onboarding.bas")

    print()
    print("Archivos generados:")
    for f in ["onboarding_reformista.xlsx", "onboarding_reformista.xlsm", "macros_onboarding.bas"]:
        p = os.path.join(out_dir, f)
        if os.path.exists(p):
            size = os.path.getsize(p)
            print(f"  📄 {f:45s} {size:>8,} bytes")

if __name__ == "__main__":
    main()
